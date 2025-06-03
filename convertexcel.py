import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
from tkinter import Tk, filedialog
import datetime
import pyodbc
import streamlit as st
import pyodbc

def get_db_connection():
    server = st.secrets["SQL_SERVER"]
    database = st.secrets["SQL_DATABASE"]
    username = st.secrets["SQL_USER"]
    password = st.secrets["SQL_PASSWORD"]
    
    conn_str = (
        f'DRIVER={{SQL Server}};'
        f'SERVER={server};'
        f'DATABASE={database};'
        f'UID={username};'
        f'PWD={password}'
    )
    return pyodbc.connect(conn_str)

def clean_cell(val):
    try:
        return str(val).replace("=", "").replace("+", "").strip()
    except:
        return val

# =====================
# STEP 1: Load customer BOM Excel
# =====================

def load_customer_bom(file_path):
    """Loads all sheets from the Excel file at a given path."""
    all_sheets = pd.read_excel(file_path, sheet_name=None)
    return all_sheets

# =====================
# STEP 2: Check required sheets
# =====================
def validate_required_sheets(all_sheets, required_sheets):
    """Validates that all required sheets are present in the loaded Excel file."""
    missing = [s for s in required_sheets if s not in all_sheets]
    if missing:
        raise ValueError(f"❌ Missing required sheets: {', '.join(missing)}")
    return all_sheets['BOM'], all_sheets['MFG']


# =====================
# STEP 3: Load company Excel template
# =====================
def load_template(path):
    """Loads the company Excel template from the specified path."""
    try:
        template_df = pd.read_excel(path)
        print("✅ Template loaded successfully.")
        return template_df
    except Exception as e:
        raise FileNotFoundError(f"❌ Failed to load template: {e}")


# =====================
# STEP 4: Validate required columns in DataFrames
# =====================
def validate_required_columns(df, name, required_cols):
    """
    Validates that required columns exist.
    Cleans the columns: strips whitespace + uppercases (for consistency).
    Checks for empty values in required columns and fails if found.
    """
    #### Check that required columns exist
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"❌ {name} is missing required columns: {', '.join(missing_cols)}")
    #### Clean columns: strip whitespace and uppercase
    for col in required_cols:
        df[col] = df[col].astype(str).str.strip().str.upper()
    
    #### Check for empty values in required columns
    empty_rows = df[df[required_cols].isin(['']).any(axis=1)] # This checks if any of the required columns have empty strings.
    if not empty_rows.empty:
        raise ValueError(f"❌ {name} has rows with missing data in required columns.")


# =====================
# STEP 5: Merge BOM and MFG DataFrames
# =====================
def merge_bom_mfg(bom_df, mfg_df, bom_keys, mfg_keys):
    merged = bom_df.merge(
        mfg_df,
        left_on=bom_keys, 
        right_on=mfg_keys, 
        how='left', 
        suffixes=('', '_MFG') # suffixes=('', '_MFG'): if any columns (like DESCRIPTION) exist in both, the MFG version gets _MFG added so nothing is overwritten.
        )
    
    # Drop any columns that start with 'ORIGINAL' except the 'ORIGINAL' column itself
    merged.drop(columns=[c for c in merged.columns if c.upper().startswith('ORIGINAL') and c != 'ORIGINAL'], inplace=True)
    return merged
################
# If multiple MFG rows match same BOM row → merge duplicates BOM row once per match.
# If BOM or MFG has duplicate keys → may duplicate merged rows.
################



# =====================
# STEP 6: Map customer columns to company template
# =====================
def map_columns_to_template(combined_df, column_mapping):
    template_df = combined_df.rename(columns=column_mapping) # Rename columns to match template
    template_df = template_df.reindex(columns=column_mapping.values()) # Reorder columns to match template
    template_df.fillna('', inplace=True) # Replace NaN with empty strings
    return template_df


def clean_output_df(df):
    return df.replace({pd.NaT: '', 'NaT': '', 'nan': ''}).fillna('')



# =====================
# STEP 7: Write filled Excel template with formatting
# =====================
def write_filled_template(template_df, bom_df, mfg_df, company_template_path, output_path, columns_to_extract):
   
    def write_dataframe_to_sheet(sheet, df, bold=True, center_headers=True):
        header_font = Font(name='Arial', size=10, bold=bold)
        regular_font = Font(name='Arial', size=10)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center' if center_headers else 'left', vertical='center')

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = regular_font
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for i, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            sheet.column_dimensions[get_column_letter(i)].width = max_len + 2

        sheet.freeze_panes = sheet['A2']

    workbook = load_workbook(company_template_path)

    desired_order = [
    'Level', 'Dwg_Item', 'Customer_Part', 'REV', 'Description_x', 'Mfr', 'MPN', 'UM', 'Unit_Qty',
    'Unit_Cost', 'PTH_Stock', 'PC_MRP_YN', 'Buyer_Cost', 'Demand', 'Last_Cost_Update', 'Notes'
    ]

    template_df = template_df[[col for col in desired_order if col in template_df.columns]]
    template_df = template_df.replace({pd.NaT: '', 'NaT': '', 'nan': ''}).fillna('')

    # -- Formatted-BOM --
    template_df = clean_output_df(template_df)
    if 'Formatted-BOM' in workbook.sheetnames:
        workbook.remove(workbook['Formatted-BOM'])
    sheet = workbook.create_sheet('Formatted-BOM')
    write_dataframe_to_sheet(sheet, template_df)

    # Apply special styles to Formatted-BOM
    bold_font = Font(name='Arial', size=10, bold=True)
    regular_font = Font(name='Arial', size=10)
    highlight_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    level_idx = template_df.columns.get_loc('Level') + 1 if 'Level' in template_df.columns else None
    desc_idx = template_df.columns.get_loc('Description') + 1 if 'Description' in template_df.columns else None
    dwg_idx = template_df.columns.get_loc('Dwg_Item') + 1 if 'Dwg_Item' in template_df.columns else None
    align_cols = ['Level', 'Dwg_Item', 'Customer_Part', 'REV', 'UM', 'Unit_Qty']
    align_indices = [template_df.columns.get_loc(c) + 1 for c in align_cols if c in template_df.columns]

    # Get the column index of 'Unit_Qty' from header row
    header_map = {sheet.cell(row=1, column=col).value.strip(): col for col in range(1, sheet.max_column + 1)}
    unit_qty_col_idx = header_map.get('Unit_Qty')
    unit_cost_col_idx = header_map.get('Unit_Cost')
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    PC_MRP_YN_col_idx = header_map.get('PC_MRP_YN')
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        level = sheet.cell(row=row[0].row, column=level_idx).value if level_idx else None
        try:
            is_lvl_0 = int(str(level).strip()) == 0
        except:
            is_lvl_0 = False


        for cell in row:
            if is_lvl_0 and unit_qty_col_idx and cell.col_idx <= unit_qty_col_idx:
                cell.fill = highlight_fill
                cell.font = bold_font
            if unit_cost_col_idx and cell.col_idx == unit_cost_col_idx:
                cell.fill = light_green_fill
                cell.alignment = center_align
                cell.number_format = '"$"  #,##0.00'
            if PC_MRP_YN_col_idx and cell.col_idx == PC_MRP_YN_col_idx:
                cell.fill = yellow_fill
            if cell.col_idx in align_indices or (is_lvl_0 and desc_idx and cell.col_idx == desc_idx):
                cell.alignment = center_align
            if dwg_idx and cell.col_idx == dwg_idx:
                cell.number_format = 'General' if is_lvl_0 else '000'
            if level_idx and cell.col_idx == level_idx and level == 0:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        sheet.row_dimensions[row[0].row].height = 13

    # -- BOM-Extract Sheet --
    if columns_to_extract:
        if 'BOM-Extract' in workbook.sheetnames:
            workbook.remove(workbook['BOM-Extract'])
        extract_df = clean_output_df(bom_df[columns_to_extract])
        extract_sheet = workbook.create_sheet('BOM-Extract')
        write_dataframe_to_sheet(extract_sheet, extract_df)

        # Highlight ITEM_NUMBER == 0
        part_idx = extract_df.columns.get_loc('PART_NUMBER') + 1 if 'PART_NUMBER' in extract_df.columns else None
        item_idx = extract_df.columns.get_loc('ITEM_NUMBER') + 1 if 'ITEM_NUMBER' in extract_df.columns else None
        blue_fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")

        for row in extract_sheet.iter_rows(min_row=2, max_row=extract_sheet.max_row):
            item_value = extract_sheet.cell(row=row[0].row, column=item_idx).value if item_idx else None
            if item_value == 0 and part_idx:
                extract_sheet.cell(row=row[0].row, column=part_idx).fill = blue_fill

        print(f"✅ Extracted columns {columns_to_extract} into 'BOM-Extract' with highlight.")

    # -- Raw-BOM Sheet --
    bom_df = clean_output_df(bom_df)
    if 'Customer BOM' in workbook.sheetnames:
        workbook.remove(workbook['Customer BOM'])
    raw_sheet = workbook.create_sheet('Customer BOM')
    write_dataframe_to_sheet(raw_sheet, bom_df)

    print("✅ Customer BOM sheet added.")

    mfg_df = clean_output_df(mfg_df)
    if 'MFG' in workbook.sheetnames:
        workbook.remove(workbook['MFG'])
    raw_sheet = workbook.create_sheet('MFG')
    write_dataframe_to_sheet(raw_sheet, mfg_df)

    print("✅ MFG sheet added.")

        # === Set desired sheet order ===
    desired_order = ["QUOTE", "Formatted-BOM", "Customer BOM", "MFG", "BOM-Extract"]

    # Ensure only existing sheets are included and preserve their order
    ordered_sheets = [sheet for name in desired_order for sheet in workbook.worksheets if sheet.title == name]
    # Add any remaining sheets not listed
    unordered_sheets = [s for s in workbook.worksheets if s not in ordered_sheets]

    # Apply final order
    workbook._sheets = ordered_sheets + unordered_sheets


    workbook.save(output_path)



# =====================
# STEP 8: Open output file (Windows only)
# =====================
def open_output_file(output_path):
    """Opens the completed Excel file using system default app (works on Windows)."""
    if os.path.exists(output_path):
        os.startfile(output_path)
    else:
        print("❌ Completed file not saved.")


# =====================
# STEP 9: Generate console summary report
# =====================
def generate_summary_report(bom_df, combined_df):
    """Prints summary report: counts, missing MPN/Mfr, etc."""
    mask = combined_df['Customer_Part'].notna() & (combined_df['Customer_Part'].astype(str).str.strip() != '')
    completed_df = combined_df[mask]


    total_rows_bom = len(bom_df)
    total_rows_output = len(completed_df)

    missing_mpn = completed_df[completed_df['MPN'].isna() | (completed_df['MPN'].astype(str).str.strip() == '')]
    missing_mfr = completed_df[completed_df['Mfr'].isna() | (completed_df['Mfr'].astype(str).str.strip() == '')]
    missing_both = completed_df[
        (completed_df['MPN'].isna() | (completed_df['MPN'].astype(str).str.strip() == '')) &
        (completed_df['Mfr'].isna() | (completed_df['Mfr'].astype(str).str.strip() == ''))
    ]

    print("\n🔎 Summary Report:")
    print(f"📄 Total rows in BOM: {total_rows_bom}")
    print(f"📦 Total rows in output: {total_rows_output}")
    print(f"✔️ Total valid parts: {len(completed_df)}")
    print(missing_mpn['Customer_Part'].tolist())
    print(f"⚠️ Missing MPN: {len(missing_mpn)}")
    print(f"⚠️ Missing Mfr: {len(missing_mfr)}")
    print(missing_mfr['Customer_Part'].tolist())
    print(f"🚨 Missing BOTH MPN and Mfr: {len(missing_both)}")
    print(missing_both['Customer_Part'].tolist())


# =====================
# MAIN WORKFLOW
# =====================

def main_process(input_path):
    required_sheets = ['BOM', 'MFG']
    company_template_path = "QUOTE_TEMPLATE_AMAT-SINGLE.xlsx"

    # Load customer BOM from provided file path
    all_sheets = load_customer_bom(input_path)
    bom_df, mfg_df = validate_required_sheets(all_sheets, required_sheets)
    template_df = load_template(company_template_path)

    # Ensure template has enough rows
    if len(bom_df) > len(template_df):
        extra = len(bom_df) - len(template_df)
        empty_rows = pd.DataFrame('', index=range(extra), columns=template_df.columns)
        template_df = pd.concat([template_df, empty_rows], ignore_index=True)

    # Determine merge keys
    if bom_df['ORIGINAL'].nunique() == 1:
        bom_keys = ['PART_NUMBER']
        mfg_keys = ['PART_NUMBER']
    else:
        bom_keys = ['ORIGINAL', 'PART_NUMBER']
        mfg_keys = ['ORIGINAL', 'PART_NUMBER']

    validate_required_columns(bom_df, 'BOM', bom_keys)
    validate_required_columns(mfg_df, 'MFG', mfg_keys)

    combined_df = merge_bom_mfg(bom_df, mfg_df, bom_keys, mfg_keys)
    template_filled_df = map_columns_to_template(combined_df, column_mapping={
        'LEVEL': 'Level',
        'ITEM_NUMBER': 'Dwg_Item',
        'PART_NUMBER': 'Customer_Part',
        'REVISION': 'REV',
        'DESCRIPTION': 'Description',
        'MANUFACTURER_NAME': 'Mfr',
        'MFG_PART_NUM': 'MPN',
        'UOM': 'UM',
        'QUANTITY': 'Unit_Qty',
        'UNIT_COST': 'Unit_Cost',
        'PTH_STOCK': 'PTH_Stock',
        'PC_MRP_FLAG': 'PC_MRP_YN',
        'NOTES': 'Notes'
    })

    # Load ERP data
    conn = get_db_connection()
    view_df = pd.read_sql("SELECT [Item Number], [Cost] FROM InventoryProductsView", conn)
    inventory_df = pd.read_sql("""
        SELECT [Item Number], [Description], [Manufacturer], [Retail],
               [Inv_LastRetailUpdate], [Special Features]
        FROM Inventory
    """, conn)
    quantity_df = pd.read_sql("SELECT [Item Number], [SumOfQuantity In Stock] FROM QtyInStock", conn)
    demand_df = pd.read_sql("""
        SELECT [Item Number], SUM([Qty Needed]) AS TotalQtyNeeded FROM (
            SELECT Kit.Kit_InvNum AS [Item Number], Kit.Kit_AllocQty AS [Qty Needed]
            FROM Kit WHERE Kit.Kit_AllocQty <> 0 AND Kit.Kit_InvNum IS NOT NULL
            UNION ALL
            SELECT Jobs.Jb_Part_Num AS [Item Number],
                   COALESCE(JobShip.Sh_Qty_Due, 0) - COALESCE(JobShip.Sh_QtyPulled, 0)
            FROM Jobs LEFT JOIN JobShip ON Jobs.Jb_Job_Num = JobShip.Sh_Job_Num
            WHERE Jobs.Jb_Part_Num IS NOT NULL
        ) AS InventoryAllocationsByPNQry GROUP BY [Item Number]
    """, conn)
    conn.close()

    for df in [view_df, inventory_df, quantity_df, demand_df]:
        df["Item Number"] = df["Item Number"].astype(str).str.strip().str.upper()
        df.drop_duplicates(subset="Item Number", keep="first", inplace=True)

    erp_combined = view_df.merge(inventory_df, on="Item Number", how="left")
    erp_combined = erp_combined.merge(quantity_df, on="Item Number", how="left")
    erp_combined = erp_combined.merge(demand_df, on="Item Number", how="left")

    erp_merged = template_filled_df.merge(erp_combined, left_on="MPN", right_on="Item Number", how="left")
    erp_merged["Unit_Cost"] = erp_merged["Retail"]
    erp_merged["PTH_Stock"] = erp_merged["SumOfQuantity In Stock"]
    erp_merged["Notes"] = erp_merged["Special Features"]
    erp_merged = erp_merged.applymap(clean_cell)
    erp_merged["Unit_Cost"] = pd.to_numeric(erp_merged["Unit_Cost"], errors="coerce")
    erp_merged = erp_merged.rename(columns={
        'Cost': 'Buyer_Cost',
        'TotalQtyNeeded': 'Demand',
        'Inv_LastRetailUpdate': 'Last_Cost_Update'
    })

    # Output path to temp file
    import tempfile
    output_path = os.path.join(tempfile.gettempdir(), "Completed_Template.xlsx")
    write_filled_template(erp_merged, bom_df, mfg_df, company_template_path, output_path, ['QUANTITY', 'PART_NUMBER', 'ITEM_NUMBER'])

    return output_path
